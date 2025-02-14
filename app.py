from flask import Flask, render_template, request
from flask_mail import Mail, Message
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pillow_heif
from PIL import Image as PILImage
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, PatternFill, Alignment


app = Flask(__name__)

# การตั้งค่าการเชื่อมต่อกับเซิร์ฟเวอร์อีเมล
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'hatta.seakh@gmail.com'
app.config['MAIL_PASSWORD'] = 'izpj wupb gxbu uysl'
app.config['MAIL_DEFAULT_SENDER'] = 'hatta.seak@gmail.com'

mail = Mail(app)

UPLOAD_FOLDER = "static/uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'heic', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ฟังก์ชันบีบอัดภาพ
def compress_image(image_path, output_path, quality=85):
    image = PILImage.open(image_path)
    image.save(output_path, "JPEG", quality=quality)

# ฟังก์ชันแปลง HEIC เป็น JPG
def convert_heic_to_jpg(heic_path):
    heif_file = pillow_heif.open(heic_path)
    image = PILImage.frombytes(heif_file.mode, heif_file.size, heif_file.data)
    jpg_path = heic_path.replace(".heic", ".jpg")
    image.save(jpg_path, format="JPEG")
    return jpg_path

# รายการตรวจสอบยานพาหนะ
important_items = {
    "ระดับน้ำหล่อเย็น , น้ำมันเครื่อง  การรั่วขณะติดเครื่องยนต์",
    "ระดับน้ำมันเบรก , คลัช , เพาเวอร์  การรั่วขณะติดเครื่องยนต์",
    "ตรวจสอบมาตรวัดบนแนวหน้าปัทม์และสัญญาณไฟเตือน (โดยเฉพาะไฟเตือนระดับน้ำมันเบรค)",
    "ตรวจสอบสภาพวาล์ว , ท่อ , ข้อต่อต่างๆ",
    "สภาพปั้มลงสินค้าและรอยการรั่วหยด",
    "เกจแสดงความร้อนเครื่องยนต์ (ดิจิตอล)"
}

inspection_items = [
    {"id": i + 1, "name": name, "important": name in important_items} for i, name in enumerate([  
        "ระดับน้ำหล่อเย็น , น้ำมันเครื่อง  การรั่วขณะติดเครื่องยนต์",
        "ระดับน้ำมันเบรก , คลัช , เพาเวอร์  การรั่วขณะติดเครื่องยนต์",
        "ระดับน้ำมันเชื้อเพลิง  การรั่วขณะติดเครื่องยนต์",
        "ทดสอบการทำงานของเครื่องยนต์และเสียงผิดปกติ",
        "ทดสอบการทำงานของระบบ  เบรก ,  ครัช  ,  เกียร์ , พวงมาลัย",
        "ตรวจสอบมาตรวัดบนแนวหน้าปัทม์และสัญญาณไฟเตือน (โดยเฉพาะไฟเตือนระดับน้ำมันเบรค)",
        "ทดสอบการทำงานใบปัดน้ำฝนและน้ำฉีดกระจก",
        "เข็มขัดนิรภัยใช้งานได้",
        "กระจกมองข้าง  ซ้าย-ขวา",
        "ระบบไฟแสงสว่างหน้า / ไฟเลี้ยว  / ไฟเบรก / ไฟหรี่ / ไฟหลังคา / แตร",
        "สภาพแบตเตอรี่และน้ำกลั่น / ฝาครอบ",
        "สภาพยางหน้า / หลัง / อะไหล่ / แรงดันลมยาง / กระทะล้อ / น็อตล้อ",
        "ถ่ายน้ำจากถังลม / วาล์วเดรนลม",
        "สภาพทั่วไปรอบรถ ( รอยเฉี่ยวชน )",
        "หมวกนิรภัย , กระบังหน้า",
        "แว่นตานิรภัย , ก๊อกเกิ้ล",
        "ถุงมือป้องกันสารเคมี",
        "ชุด PVC ป้องกันสารเคมี",
        "รองเท้านิรภัย , รองเท้าบู๊ทยาง",
        "หน้ากากกรองสารเคมี",
        "ขวดน้ำล้างตา",
        "เข็มขัดกันตกจากที่สูง",
        "ป้ายสัญลักษณ์แสดงข้อมูลสินค้า(แผ่น UN.ใหญ่) ทั้งหมด 3 ด้าน",
        "เอกสารความปลอดภัยเกี่ยวกับสารเคมี (SDS) และแผนฉุกเฉิน",
        "ถังดับเพลิง ( เกจอยู่ในพื้นที่สีเขียว )",
        "ไม้หมอนหนุนล้อ",
        "กรวยจราจร 2 อัน",
        "ถังเดรนประจำรถ ( ถังสแตนเลสสำหรับรถโซลเวนท์ )",
               "ลิ่ม, ค้อน",
        "พลั่ว",
        "เสื้อกั๊กสะท้อนแสง",
        "วัสดุสำหรับซับสาร",
        "ปูนขาว",
        "แผ่นพลาสติก",
        "เทปกั้นบริเวณ",
        "ไฟฉาย",
        "ประแจรวมหลายเบอร์ 1 ชุด",
        "ตรวจสภาพแท้งค์หารอยรั่มซึม",
        "ตรวจสอบสภาพวาล์ว , ท่อ , ข้อต่อต่างๆ",
        "หน้าแปลน 2 นิ้ว",
        "หน้าแปลน 3 นิ้ว",
        "ล๊อกเร็วตัวผู้ 2 นิ้ว ( Part E )",
        "ล๊อกเร็วตัวเมีย 2 นิ้ว ( Part C )",
        "สภาพสายส่งสินค้า",
        "สภาพสายสำรองส่งสินค้า ( ถ้ามี )",
        "สภาพปั้มลงสินค้าและรอยการรั่วหยด",
        "สภาพสวิทช์ปุ่มกดฉุกเฉินรอบรถ 3 จุด และทดสอบการทำงาน",
        "ทดสอบการทำงานของวาล์วฉุกเฉิน ( Internal Valve )",
        "ปลั๊ก 5 ขา 16 แอมป์",
        "ปลั๊ก 5 ขา 32 แอมป์",
        "เกจแสดงความร้อนเครื่องยนต์ (ดิจิตอล)",
        "ข้อต่อเกลียวนอกละเอียดเกลียวในหยาบ 3 นิ้ว",
        "ข้อต่อเกลียวนอกหยาบเกลียวในละเอียด 3 นิ้ว",
        "ข้อต่อตัวผู้ 3 นิ้ว เกลียวใน(Part A)",
        "ข้อต่อตัวเมีย 3 นิ้ว เกลียวอก(Part B)",
        "หัวกรอกสำหรับกรอกสารเคมี"
    ])
]

# ฟังก์ชันส่งอีเมล
def send_email_with_attachment(excel_filename):
    msg = Message(
        'Inspection Report',  
        recipients=['hatta.seak@gmail.com', 'songdee.eng@songdeegps.com']
    )
    with app.open_resource(excel_filename) as fp:
        msg.attach(excel_filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', fp.read())

    try:
        mail.send(msg)
    except Exception as e:
        print(f"Error sending email: {e}")

# หน้าแรกของแอปพลิเคชัน
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        license_plate = request.form['license_plate']
        date = request.form['date']
        driver = request.form['driver']
        inspection_data = []

        # เก็บข้อมูลการตรวจสอบ
        for item in inspection_items:
            status = request.form.get(f"status_{item['id']}")
            image = request.files.get(f"image_{item['id']}")

            if image and allowed_file(image.filename):
                filename = secure_filename(image.filename)  # ใช้ secure_filename ที่นำเข้ามา
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                image.save(filepath)

                # บีบอัดภาพก่อนบันทึก
                compress_image(filepath, filepath)

                if filename.lower().endswith(".heic"):
                    convert_heic_to_jpg(filepath)

            inspection_data.append({
                'id': item['id'],
                'name': item['name'],
                'status': status,
                'image': filepath if image else None
            })

        # สร้างไฟล์ Excel
        excel_filename = generate_excel_report(inspection_data, license_plate, date, driver)

        # ส่งอีเมล
        send_email_with_attachment(excel_filename)

        return render_template('index.html', items=inspection_items, message="ส่งรายงานแล้ว")
    return render_template('index.html', items=inspection_items)

# สร้างไฟล์ Excel
def generate_excel_report(data, license_plate, date, driver):
    df = pd.DataFrame(data)

    excel_filename = f"static/{license_plate}_inspection_report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        workbook = writer.book
        sheet = workbook.create_sheet('Inspection Report')

        # ตั้งค่าหัวข้อ
        sheet['A1'], sheet['B1'] = 'License Plate', license_plate
        sheet['A2'], sheet['B2'] = 'Date', date
        sheet['A3'], sheet['B3'] = 'Driver', driver

        # ตั้งค่าฟอนต์หัวข้อ
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        
        for col in ['A', 'B', 'C']:
            for row in range(1, 4):
                sheet[f'{col}{row}'].font = header_font
                sheet[f'{col}{row}'].fill = header_fill
                sheet[f'{col}{row}'].alignment = Alignment(horizontal='center')

        # เขียนหัวตาราง
        sheet.append(["Inspection Item", "Status", "Image"])
        
        row_index = 5  # เริ่มเขียนข้อมูลจากแถวที่ 5

        for item in data:
            # สีสำหรับรายการที่สำคัญ
            important_fill = PatternFill(start_color="FFEB9C", fill_type="solid")  # สีเหลือง

            # ตรวจสอบว่ารายการนี้สำคัญไหม และมาร์คสี
            if item['name'] in important_items:
                sheet.append([item['name'], item['status']])
                sheet.row_dimensions[row_index].height = 100  # ปรับความสูงของแถวให้พอดีกับรูปภาพ
                # มาร์คสีแถวสำหรับรายการสำคัญ
                for col in ['A', 'B', 'C']:
                    sheet[f'{col}{row_index}'].fill = important_fill

                if item['image']:
                    try:
                        img = Image(item['image'])
                        img.width, img.height = 150, 100  # ปรับขนาดรูปภาพ
                        img_anchor = f"C{row_index}"  # ใส่รูปในคอลัมน์ C
                        sheet.add_image(img, img_anchor)
                    except Exception as e:
                        print(f"Error inserting image: {e}")
            else:
                sheet.append([item['name'], item['status']])
                sheet.row_dimensions[row_index].height = 100  # ปรับความสูงของแถวให้พอดีกับรูปภาพ
                if item['image']:
                    try:
                        img = Image(item['image'])
                        img.width, img.height = 150, 100  # ปรับขนาดรูปภาพ
                        img_anchor = f"C{row_index}"  # ใส่รูปในคอลัมน์ C
                        sheet.add_image(img, img_anchor)
                    except Exception as e:
                        print(f"Error inserting image: {e}")

            row_index += 1  # ขยับไปแถวถัดไป

        # ตั้งค่าความกว้างของคอลัมน์
        sheet.column_dimensions['A'].width = 50  # รายการตรวจสอบ
        sheet.column_dimensions['B'].width = 20  # สถานะ
        sheet.column_dimensions['C'].width = 30  # รูปภาพ

    return excel_filename


if __name__ == '__main__':
    app.run(debug=True)